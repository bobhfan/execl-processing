import openpyxl
from datetime import datetime


def read_data(file_info_dict):

    file_name = file_info_dict['file_name']
    interest_field_list = file_info_dict['interest_field_list']
    header_row = file_info_dict['header_row']
    first_key_name = file_info_dict['first_key_name']
    second_key_name = file_info_dict['second_key_name']
    datetime_fmt = file_info_dict['datetime_fmt']
    dict_level = file_info_dict['dict_level']

    wb= openpyxl.load_workbook(file_name, data_only = True)
    print(wb.sheetnames)
    sheet = wb.active

    field_name_dict = {}

    for x in range (header_row,header_row + 1):
        for y in range(1,sheet.max_column + 1):
            field_name_dict.update({y:sheet.cell(row=x,column=y).value})

    print("Ther are {} rows in {} sheet of {} ".format(sheet.max_row, wb.sheetnames[0], file_name ))

    if 'seq' == second_key_name:
        seq_no = 1

    second_key = ''
    data_dict = {}
    for line in range(header_row + 1, sheet.max_row+1):
    # for line in range(3, 155):
        tmp_dict = {}
        for colum_n in range(1, sheet.max_column+1):
            # This is for extract item list info to create the dict for the first step
            # {item_no : { interesting_field : value}}
            if field_name_dict[colum_n]:
                if first_key_name == field_name_dict[colum_n]:
                    if isinstance(sheet.cell(row=line, column=colum_n).value, float):
                        first_key = str(int(sheet.cell(row=line, column=colum_n).value))
                    else:
                        first_key = sheet.cell(row=line, column=colum_n).value

                elif second_key_name and second_key_name == field_name_dict[colum_n]:
                    if isinstance(sheet.cell(row=line, column=colum_n).value, float):
                        second_key = str(int(sheet.cell(row=line, column=colum_n).value))
                    else:
                        second_key = sheet.cell(row=line, column=colum_n).value                

                elif field_name_dict[colum_n] in interest_field_list:
                    if isinstance(sheet.cell(row=line, column=colum_n).value, datetime):
                        date_value = sheet.cell(row=line, column=colum_n).value.strftime(datetime_fmt)
                        tmp_dict.update({field_name_dict[colum_n] : date_value})
                    else:
                        tmp_dict.update({field_name_dict[colum_n] : sheet.cell(row=line, column=colum_n).value})

        if 'seq' == second_key_name:
            seq_no += 1
        # if first_key == 'PAF CO008':
        #     print('got it')
        if 'seq' == second_key_name:
            data_dict[first_key + '---' + str(seq_no)] = tmp_dict
        elif second_key_name and 'seq' != second_key_name and second_key and dict_level == 1:
            data_dict[str(second_key) + '---' + str(first_key) ] = tmp_dict
        elif dict_level == 2:
            if not second_key:
                second_key = 'unknown'
                print("{} err: empty record for {} ".format(file_name, first_key))
            if first_key not in data_dict:
                data_dict[first_key] = {}

            data_dict[first_key][second_key] = tmp_dict
        else:
            data_dict[first_key] = tmp_dict

    return data_dict