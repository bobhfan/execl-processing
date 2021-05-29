import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import PatternFill, Alignment
from datetime import datetime, timedelta
from datetime import date as date_op
import pprint
import argparse
import json
from collections import OrderedDict, defaultdict
import holidays
import sys
import os


from read_data import read_data
from file_info import File_Info_Dict

FMT = '%Y-%m-%d'
YEAR_FMT = '%Y'

# time = datetime.strptime("2021-02-13 00:00:00", FMT)

# date_time = datetime.now()

# print("now is {}".format(date_time.strftime(FMT)))

WEEK_DAY = {
    0 : 'Mon',
    1 : "Tue",
    2 : "Wen",
    3 : "Thu",
    4 : "Fri",
    5 : "Sat",
    6 : "Sun",
}


Interest_Duty_List = ["FILM", "BAG",  "SEASONING"]

Valid_Purchase_Period = timedelta(days=365)
Current_day = datetime.today() + timedelta(days=-1)


# append the detail info of each BOM item extract from item list_step_1.json to generate BOM List_step_2.json file
# Expand the Customer order list item to all the BOM items involved into and filter the info based on the Interest_Duty_Class
def step_2_processsing(dict_step_2):

    bom_dict = {}
    for file_name in dict_step_2.keys():
        if "bom" in file_name:
            bom_name = file_name
        elif "item" in file_name:
            item_name = file_name
        elif "purchase" in file_name:
            purchase_name = file_name
        elif "customer" in file_name:
            customer_name = file_name
        with open(file_name) as json_file:
            dict_step_2[file_name] = json.load(json_file)

    purchase_dict = {}

    seq_no = 0
    for item_no_seq, internal_dict in dict_step_2[purchase_name].items():
        item_no, _ = item_no_seq.split('---')
        if item_no not in purchase_dict:
            purchase_dict[item_no] = {}
            purchase_dict[item_no]['promise'] = {}
            purchase_dict[item_no]['request'] = {}

        if internal_dict['Promised Receipt Date']:
            purchase_dict[item_no]['promise'][internal_dict['Promised Receipt Date'] + '---' + str(seq_no)] = \
                                            internal_dict['Outstanding Quantity']
        elif internal_dict['Requested Receipt Date']:
            purchase_dict[item_no]['request'][internal_dict['Requested Receipt Date'] + '---' + str(seq_no)] = \
                                            internal_dict['Outstanding Quantity']

        seq_no += 1        


    for item_no, top_dict in purchase_dict.items():
        tmp_dict = OrderedDict(sorted(top_dict['promise'].items()))
        purchase_dict[item_no] = {}
        purchase_dict[item_no]['promise'] = tmp_dict
        tmp_dict = OrderedDict(sorted(top_dict['request'].items()))
        purchase_dict[item_no]['request'] = tmp_dict


    with open(purchase_name.replace('_1.j', '_2.j'), "w") as json_file:
        json.dump(purchase_dict, json_file, indent = 4)

    bom_dict = dict_step_2[bom_name]
    filtered_dict = {}

    for key, value in dict_step_2[customer_name].items():
        customer_document_no, bom_no = key.split('---')
        customer_dict = value
        # search the bom_no in the item list to determine if there is some storage for this to compensate the quantity
        if bom_no in dict_step_2[item_name]:
            if 'FG' in dict_step_2[item_name][bom_no].get("Item Category Code", 'unknown'):
                on_hand_quantity = dict_step_2[item_name][bom_no].get("Quantity on Hand", 0.0)
                dict_step_2[customer_name][key].update({"Quantity on Hand from Item" : on_hand_quantity})

        if bom_no in bom_dict:
            # print("there is associatekey in BOM for {}".format(value['Document No.']))
            bom_items_dict = bom_dict.get(bom_no, 'not found')

            for bom_item_no, detail_dict in bom_items_dict.items():

                value.update({bom_item_no:detail_dict})

            filtered_dict.update({key:value})
        else:
            print("customer order list err: {} no associate entry in BOM".format(bom_no))

    with open(customer_name.replace('_1.j', '_2.j'), "w") as json_file:
        json.dump(filtered_dict, json_file, indent = 4)

def step_2_1_processing(working_dir):
    with open(working_dir+'customer order list_step_2.json') as json_file:
        custom_dict = json.load(json_file)

    tmp_dict = {}
    seq_no = 1

    bom_no_list =[combined_key.split('---')[-1] for combined_key in custom_dict.keys()]

    bom_no_no_duplicate_list = list(set(bom_no_list))

    # find all order with same bom_no and sorted by ship date
    for key in bom_no_no_duplicate_list:
        tmp_dict[key] = []
        for dict_key, top_dict in custom_dict.items():
            ship_date = top_dict["Shipment Date"]
            if key == dict_key.split('---')[-1]:
                customer_no, bom_no = dict_key.split('---')
                tmp_dict[key].append('---'.join((bom_no, ship_date, customer_no)))
        
        tmp_dict[key].sort()

    tmp_2_dict = {}
    for bom_no, combinde_key_list in tmp_dict.items():
        tmp_2_dict[bom_no] = []
        for combined_key in combinde_key_list:
            bom_no_1, ship_date, customer_no = combined_key.split('---')
            # if bom_no == bom_no_1:
            search_key = '---'.join((customer_no, bom_no))
            order_quantiry = custom_dict[search_key]['Quantity']
            quantity_on_hand = custom_dict[search_key].get('Quantity on Hand from Item', 0)
            tmp_2_dict[bom_no].append((combined_key, order_quantiry, quantity_on_hand))

    tmp_3_dict = {}
    for bom_no, entries in tmp_2_dict.items():

        sum_of_quantity = 0
        all_clear = False
        tmp_list = []
        for index in range(len(entries)):
            sum_of_quantity += entries[index][1]
            if sum_of_quantity >= entries[0][-1]:
                rest_quantity = sum_of_quantity - entries[0][-1]
                break
            if index == len(entries) - 1 and sum_of_quantity <= entries[0][-1]:
                rest_quantity = sum_of_quantity - entries[0][-1]
                all_clear = True
                print("bom_no: {} all of element need be changed to zero, and still rest of {}".format(bom_no, rest_quantity))
        
        if index == 0 and len(entries) == 1 and all_clear == False:
            tmp_list.append((entries[0][0], rest_quantity, 0))
        elif all_clear == False:
            print("bom_no: {} only [0 : {}] need be changed and the rest is {}".format(bom_no, index, rest_quantity))
            for cnt in range(index):
                tmp_list.append((entries[cnt][0], 0, 0))
            tmp_list.append((entries[index][0], rest_quantity, 0))
            for cnt in range(index+1, len(entries)):
                tmp_list.append((entries[cnt][0], entries[cnt][1], 0))

        if all_clear:
            for cnt in range(len(entries)):
                tmp_list.append((entries[cnt][0], 0, 0))

        tmp_3_dict[bom_no] = tmp_list
            

    for key in tmp_2_dict.keys():
        if "PI 05350" == key:
            print(tmp_3_dict[key])
            print(tmp_2_dict[key])
        if not (len(tmp_2_dict[key]) == len(tmp_3_dict[key])) :
            print("{}'s entries {} not equal {}".format(key, len(tmp_2_dict[key]), len(tmp_3_dict[key]) ))

        for entry in tmp_3_dict[key]:
            if entry[1] < 0:
                print("{} demand < 0".format(key))

    with open(working_dir+'customer order list_step_2_1_interm.json', "w") as json_file:
        json.dump(tmp_3_dict, json_file, indent = 4)

    # ley's compensate the quantiry of each customer order
    for bom_no in tmp_3_dict.keys():
        for entry in tmp_3_dict[bom_no]:
            bom_no, shipment_date, customer_no = entry[0].split('---')
            modified_quantity = entry[1]
            modified_on_hand_quantity = entry[2]

            if custom_dict['---'.join((customer_no,bom_no))]['Shipment Date'] == shipment_date:
                custom_dict['---'.join((customer_no,bom_no))]['Outstanding Quantity'] = str(modified_quantity) + '---' + 'Mod'
                custom_dict['---'.join((customer_no,bom_no))]['Quantity on Hand from Item'] = modified_on_hand_quantity

            else :
                print('something error with record of {}'.format('---'.join((customer_no,bom_no))))

    with open(working_dir+'customer order list_step_3.json', "w") as json_file:
        json.dump(custom_dict, json_file, indent = 4)

# use bom_no_id in the BOM List as the key to reorganize the customer order list to gather all customer order which belong to
# one bom item
def step_3_processing(file_name_step_3_dict):

    with open(file_name_step_3_dict['customer_name']) as json_file:
        custom_dict = json.load(json_file)

    final_dict = {}

    for custom_no, top_dict in custom_dict.items():

        # first iteration: search item dict
        for item_no, level_2_dict in top_dict.items():
            if isinstance(level_2_dict, dict):
                tmp_dict = {}
                real_key = item_no

                # if materail not exist, create empty dict for adding
                if real_key not in final_dict:
                    final_dict.update({real_key:{}})

                # add bom detail info into dict
                for key, value in level_2_dict.items():
                    if "Description" in key :
                        final_dict[real_key].update({key:value})

                tmp_dict.update({"Production Quantity": level_2_dict["Quantity"]})
                tmp_dict.update({"Scrap %": level_2_dict["Scrap %"]})

                # after add bom info into the dict, should append customer order info
                final_dict[real_key].update({custom_no:tmp_dict})
                # append customer order info to each item entry
                for key, value in top_dict.items():
                    if not isinstance(value, dict):
                        final_dict[real_key][custom_no].update({key:value})


    with open(file_name_step_3_dict['item_name']) as json_file:
        item_dict = json.load(json_file)

    # append item info at top_dict for each item_no
    # find all items not showingup in BOM list but exist in the item list
    bom_item_list = []
    for bom_no, top_dict in final_dict.items():
        if bom_no in item_dict:
            for item_detail_key, value in item_dict[bom_no].items():
                top_dict[item_detail_key] = value
            bom_item_list.append(bom_no)

    for key, top_dict in item_dict.items():
        if key not in bom_item_list:
            final_dict[key] = top_dict


        # print("{} of duty '{}' showd up in {} customer orders".format(bom_key, top_dict['Duty Class'], order_cnt))

    with open(file_name_step_3_dict['customer_name'].replace('_3.j', '_4.j'), "w") as json_file:
        json.dump(final_dict, json_file, indent = 4)

# print("now is {}".format(date_time.strftime(FMT)))


def generate_puchase_data(input_dict):
    exist_date_list = []
    output_dict = {}
    for date_seq, quantity in input_dict.items():
        date, _ = date_seq.split('---')

        if date not in exist_date_list:
            exist_date_list.append(date)

            output_dict[date] = quantity
        else:
            output_dict[date] += quantity

    return output_dict

def distribute_data_to_different_period(input_dict, days_list, future_4_weeks_day_list):

    single_days_list = days_list[0:-4]
    output_dict = {}

    output_dict = defaultdict(lambda: float(0))
    # extra processing to accumulate the demand in the future 4 weeks
    for timestamp in input_dict.keys():
        if timestamp in future_4_weeks_day_list:
            time_diff = datetime.strptime(timestamp, FMT).date() - datetime.strptime(future_4_weeks_day_list[0], FMT).date()
            if timedelta(days=0) <= time_diff <= timedelta(days=6):
                output_dict[future_4_weeks_day_list[6]] += input_dict[timestamp]

            elif timedelta(days=6) < time_diff <= timedelta(days=13):
                output_dict[future_4_weeks_day_list[13]] += input_dict[timestamp]
            elif timedelta(days=13) < time_diff <= timedelta(days=20):
                output_dict[future_4_weeks_day_list[20]] += input_dict[timestamp]
            elif timedelta(days=20) < time_diff <= timedelta(days=27):
                output_dict[future_4_weeks_day_list[27]] += input_dict[timestamp]
        elif timestamp not in single_days_list:
            output_dict['Past Due'] += input_dict[timestamp]
        else:
            output_dict[timestamp] = input_dict[timestamp]

    return output_dict



def write_to_xls_file(file_name_step_4_dict):
    final_order_dict = OrderedDict()
    with open(file_name_step_4_dict['customer_name']) as json_file:
        final_order_dict = json.load(json_file)

    with open(file_name_step_4_dict['genpak_name']) as json_file:
        genpak_name_dict = json.load(json_file)

    with open(file_name_step_4_dict['superpufft']) as json_file:
        superpufft_dict = json.load(json_file)

    genpak_field_list = ["GP OH", "WIP", "On Order", "Target Date", "Ship Qy", "Del Date"]
    supperpufft_field_list = ["Current Stock/KG", "Next Available date & Quantity/KG"]

    book = openpyxl.Workbook()
    sheet = book.active

    row_1_list = [  "Vendor",
                    'Production Code',
                    "Duty Class",
                    'Description',
                    "Base Unit of Measure",
                    "Total Demand",
                    "Key Description",
                    "Past Due",
                ]

    specific_column_list = [
                        'On Hand (W1)',
                        'On Hand (W2)',
                        'Vendor Floor Stock',
                        'All Open PO & BPO',
                        'Purchases Receipts',
                        'Demand',
                        'Available Inventory',
                        'Total Ending Balance',
    ]

    future_4_weeks_day_list = []

    header_font = Font(color="000000")
    sub_row_font = [
                        Font(color="000000"),
                        Font(color="000000"),
                        Font(color="000000"),
                        Font(color="000000"),
                        Font(color="000000"),
                        Font(color="000000"),
                        Font(color="000000"),
                        Font(color="000000"),
                    ]


    for offset in range(14):
        next_day = Current_day + timedelta(days=offset+1)
        row_1_list.append(next_day.strftime(FMT))

    for offset_days in range(1, 29):
        if (offset_days % 7) == 0:
            next_week_day = next_day + timedelta(days=offset_days)
            row_1_list.append(next_week_day.strftime(FMT))
        future_4_weeks_day_list.append((next_day + timedelta(days=offset_days)).strftime(FMT))

    print("row list is {}".format(row_1_list))

    row_cnt = 1
    column_cnt = 1

    for column_cnt in range(len(row_1_list)):
        sheet.cell(row=row_cnt, column=column_cnt+1).value = row_1_list[column_cnt]
        sheet.cell(row=row_cnt, column=column_cnt+1).font = header_font
        sheet.cell(row=row_cnt, column=column_cnt+1).alignment = Alignment(horizontal='center')

        if column_cnt > len(row_1_list) - 5 :
            sheet.cell(row=row_cnt, column=column_cnt+1).fill = PatternFill(start_color="CDD1D6", end_color="CDD1D6", fill_type = "solid")
        try:
            date = datetime.strptime(row_1_list[column_cnt], FMT)
            week_day = WEEK_DAY[date.weekday()]
            sheet.cell(row=row_cnt, column=column_cnt+1).value += ('\n' + week_day)

            year =date.strftime(YEAR_FMT)

            if date.date() in holidays.Canada(years = int(year)).keys() or date.weekday() > 4:
                sheet.cell(row=row_cnt, column=column_cnt+1).fill = PatternFill(start_color="B3F2FF", end_color="B3F2FF", fill_type = "solid")
            elif date.date() in holidays.Canada(years = int(year)).keys():
                sheet.cell(row=row_cnt, column=column_cnt+1).value += ('\n' + 'holiday')
        except:
            sheet.cell(row=row_cnt, column=column_cnt+1).fill = PatternFill(start_color="B3F2FF", end_color="B3F2FF", fill_type = "solid")

    sheet.freeze_panes = "A2"

    sheet.print_title_rows='1:1'

    row_cnt += 1
    column_cnt = 1

    with open(file_name_step_4_dict['purchase_name']) as json_file:
        purchase_dict = json.load(json_file)



    modified_customer_order_dict = {}
    for bom_key, top_dict in final_order_dict.items():

        modified_customer_order_dict[bom_key] = []

        if bom_key == "RC 077" or bom_key == "RA 254":
            print("Got RA 263 or 254")

        if top_dict.get("Duty Class", "unknown") in Interest_Duty_List:

            demand_dict = defaultdict(lambda: float(0))
            demand_dict_tmp = defaultdict(lambda: float(0))


            purchase_promise_dict = {}
            purchase_request_dict = {}

            if bom_key in purchase_dict:

                purchase_promise_intermediate_dict = generate_puchase_data(purchase_dict[bom_key]['promise'])
                purchase_promise_dict = distribute_data_to_different_period(purchase_promise_intermediate_dict, row_1_list, future_4_weeks_day_list)

                purchase_request_intermediate_dict = generate_puchase_data(purchase_dict[bom_key]['request'])
                purchase_request_dict = distribute_data_to_different_period(purchase_request_intermediate_dict, row_1_list, future_4_weeks_day_list)


            for middle_key, middle_dict in top_dict.items():
                if isinstance(middle_dict, dict):
                    per_unit_production_quantity = float(middle_dict["Production Quantity"])
                    scrap = float(middle_dict["Scrap %"])
                    # todo@ add color to identify modified
                    if '---Mod' in middle_dict["Outstanding Quantity"]:
                        order_quantity = float(middle_dict["Outstanding Quantity"].replace('---Mod', ''))
                        # add a dict to storage the modified quantity for this bom
                        modified_customer_order_dict[bom_key].append(middle_dict["Shipment Date"])

                    demand_quantity = order_quantity * per_unit_production_quantity * (1.0 + scrap * 0.01)

                    # till now only occur once
                    if middle_dict["Shipment Date"] not in demand_dict_tmp:
                        demand_dict_tmp.update({middle_dict["Shipment Date"]: demand_quantity})
                    # already have one, accumulate the demand on it
                    else:
                        demand_sum = demand_dict_tmp[middle_dict["Shipment Date"]] + demand_quantity
                        demand_dict_tmp.update({middle_dict["Shipment Date"]: demand_sum})

            demand_dict = distribute_data_to_different_period(demand_dict_tmp, row_1_list, future_4_weeks_day_list)

            # here put each value into cell
            # loop for 8 sub_row for each item no.
            for sub_row in range(8):
                sheet.cell(row=row_cnt, column=1).value = top_dict.get("Vendor No.", "unknown")
                sheet.cell(row=row_cnt, column=2).value = bom_key
                sheet.cell(row=row_cnt, column=3).value = top_dict.get("Duty Class", "uknown")
                sheet.cell(row=row_cnt, column=4).value = top_dict.get('Description', "uknown")
                sheet.cell(row=row_cnt, column=5).value = top_dict.get("Base Unit of Measure", "uknown")

                sheet.cell(row=row_cnt, column=7).value = specific_column_list[sub_row]
                sheet.cell(row=row_cnt, column=7).font = sub_row_font[sub_row]

                total_demand = 0


                # add demand number into the excel
                if 'Demand' in specific_column_list[sub_row]:
                    # put demand value for each following 14 days
                    for column_no in range(7, len(row_1_list)):
                        if row_1_list[column_no] in demand_dict:
                            sheet.cell(row=row_cnt, column=column_no + 1).value = int(round(demand_dict[row_1_list[column_no]]))

                            # change the demand be modified by on hand storage to another color
                            if row_1_list[column_no] in modified_customer_order_dict[bom_key]:
                                sheet.cell(row=row_cnt, column=column_no + 1).font = Font( color='B30000', bold=True, italic=True)

                            total_demand += sheet.cell(row=row_cnt, column=column_no + 1).value

                elif 'All Open PO & BPO'in specific_column_list[sub_row]:
                    # put demand value for each following 14 days
                    for column_no in range(7, len(row_1_list)):
                        if row_1_list[column_no] in purchase_request_dict:
                            sheet.cell(row=row_cnt, column=column_no + 1).value = int(round(purchase_request_dict[row_1_list[column_no]]))

                elif 'Purchases Receipts'in specific_column_list[sub_row]:
                    # put demand value for each following 14 days
                    for column_no in range(7, len(row_1_list)):
                        if row_1_list[column_no] in purchase_promise_dict:
                            sheet.cell(row=row_cnt, column=column_no + 1).value = int(round(purchase_promise_dict[row_1_list[column_no]]))
                            sheet.cell(row=row_cnt, column=column_no + 1).font = Font( color='00004D', bold=True, italic=True)

                elif 'On Hand (W1)' in specific_column_list[sub_row]:
                    sheet.cell(row=row_cnt, column=8).value = int(top_dict.get("W1", 0))

                elif 'On Hand (W2)' in specific_column_list[sub_row]:
                    sheet.cell(row=row_cnt, column=8).value = int(top_dict.get("W2", 0))

                elif 'Vendor Floor Stock' in specific_column_list[sub_row]:
                    column_no = 8
                    if bom_key in genpak_name_dict:
                        for index in range(len(genpak_field_list)):
                            sheet.cell(row=row_cnt, column=column_no+index).value = genpak_name_dict[bom_key][genpak_field_list[index]]
                    elif bom_key in superpufft_dict:
                        for index in range(len(supperpufft_field_list)):
                            sheet.cell(row=row_cnt, column=column_no+index).value = superpufft_dict[bom_key][supperpufft_field_list[index]]


                sheet.cell(row=row_cnt, column=6).value = int(round(total_demand))

                row_cnt += 1
                column_cnt = 1

            # after all value could be read from the dict directly, and prepared moving to nenext iteration for next bmo_no, we could start calculating
            # some cell based on the cell value
            day_column_start = 9
            past_due_column = 8
            previous_w1_past_due_row = row_cnt - 8


            previous_purchase_receipts_row = previous_w1_past_due_row + 4
            previous_demand_row = previous_w1_past_due_row + 5
            previous_available_inventory_row = previous_w1_past_due_row + 6
            previous_total_ending_balance_row = previous_w1_past_due_row + 7
            
            # cell(row=Available Inventory, column=Past due) ---> W1 + purchasde Receipts 
            if not sheet.cell(row=previous_w1_past_due_row, column=past_due_column).value:
                sheet.cell(row=previous_w1_past_due_row, column=past_due_column).value = 0
            if not sheet.cell(row=previous_purchase_receipts_row, column=past_due_column).value:
                sheet.cell(row=previous_purchase_receipts_row, column=past_due_column).value = 0
            if not sheet.cell(row=previous_demand_row, column=past_due_column).value:
                sheet.cell(row=previous_demand_row, column=past_due_column).value = 0


            sheet.cell(row=previous_available_inventory_row, column=past_due_column).value = \
                int(sheet.cell(row=previous_w1_past_due_row, column=past_due_column).value) - \
                int(sheet.cell(row=previous_demand_row, column=past_due_column).value)

            # cell(row=Total Ending balance,column=past Due) ---> W1 + purchasde Receipts - demand
            sheet.cell(row=previous_total_ending_balance_row, column=past_due_column).value = \
                int(sheet.cell(row=previous_available_inventory_row, column=past_due_column).value) + \
                int(sheet.cell(row=previous_purchase_receipts_row, column=past_due_column).value)            

            for last_2_row_column_no in range(9, len(row_1_list)):
                if not sheet.cell(row=previous_demand_row, column=last_2_row_column_no).value:
                    sheet.cell(row=previous_demand_row, column=last_2_row_column_no).value = 0
                if not sheet.cell(row=previous_purchase_receipts_row, column=last_2_row_column_no).value:
                    sheet.cell(row=previous_purchase_receipts_row, column=last_2_row_column_no).value = 0

                sheet.cell(row=previous_available_inventory_row, column=last_2_row_column_no).value = \
                    int(sheet.cell(row=previous_available_inventory_row, column=last_2_row_column_no - 1).value) - \
                    int(sheet.cell(row=previous_demand_row, column=last_2_row_column_no).value)

                sheet.cell(row=previous_total_ending_balance_row, column=last_2_row_column_no).value = \
                    int(sheet.cell(row=previous_available_inventory_row, column=last_2_row_column_no).value) + \
                    int(sheet.cell(row=previous_purchase_receipts_row, column=last_2_row_column_no).value)

        column_cnt = 1

    print("write {} row".format(row_cnt))


    book.save("output_{}.xlsx".format(datetime.now().strftime(FMT)))


def build_parser():
    '''
    Build command line options.
    '''
    parser = argparse.ArgumentParser(
        description='Procesing the xls files')

    # Positional Arguments
    parser.add_argument("-d", "--dst_dir",
                        type=str,
                        help='destination direcoty')

    parser.add_argument("-r", "--running_steps",
                        type=str,
                        help='all steps or skip first reading which consume most of time <all|skip>')

    return parser

def main():
    '''
    Args:
        argv[1](str) :  destination directory
        argv[2](str) :  step all or exclude step_1 of reading inital execl file

    Returns:
    '''


    # Load argparse and parse arguments.
    cmd_parser = build_parser()
    args = cmd_parser.parse_args(sys.argv[1:])


    #Commnd from Zabbix server to query the RPD stats
    # working_dir = args.dst_dir
    working_dir = './'

    data_files = [(x[0], x[2]) for x in os.walk(working_dir)]
    for path_files in data_files:
        for file_name in path_files[1]:
            if '.xlsx' in file_name and 'Zone' not in file_name:
                print(file_name.lower())
                create_date = datetime.fromtimestamp(os.stat(working_dir+file_name).st_ctime).date()
                if datetime.now().date() == create_date:
                    print("Create on same day, continue")
                    os.rename(working_dir+file_name, working_dir+file_name.lower())
                else:
                    print("{} created on different day, exit please double check".format(working_dir+file_name))

    # step_2 running some data manipulation on the step_1 json files
    step_2_file_dict = {}
    for file_dict in File_Info_Dict:
        file_name = file_dict['file_name'].replace('.xlsx', '_step_1.json')
        step_2_file_dict[file_name] = {}

    file_name_step_3_dict = {
                            'customer_name' : working_dir + 'customer order list_step_3.json',
                            'item_name' : working_dir + 'item list_step_1.json',
    }

    file_name_step_4_dict = {
                            'customer_name' : working_dir + 'customer order list_step_4.json',
                            'purchase_name' : working_dir + 'purchase lines_step_2.json',
                            'genpak_name' : working_dir + 'genpak releases - current_step_1.json',
                            'superpufft' : working_dir + 'superpufft 2020 pricing  inventory report may 27, 2021_step_1.json',
    }

    if 'all' in args.running_steps:

        # step_1 reading all data into json file
        for file_dict in File_Info_Dict:
            file_name = file_dict['file_name']
            info_dict = read_data(file_dict)
            with open(file_name.replace('.xlsx', '_step_1.json'), "w") as json_file:
                json.dump(info_dict, json_file, indent = 4) 
            if 'bom' not in file_name:
                print("{} record in {}".format(len(info_dict), file_name))
            else:
                total_item_no = 0
                for key, value in info_dict.items():
                    total_item_no += len(value)

                print("{} record in the {}".format(total_item_no, file_name))


    step_2_processsing(step_2_file_dict)

    step_2_1_processing(working_dir)

    step_3_processing(file_name_step_3_dict)

    write_to_xls_file(file_name_step_4_dict)

if __name__ == '__main__':
    main()
