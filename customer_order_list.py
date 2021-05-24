import openpyxl
from datetime import datetime, timedelta
import pprint
import json

FMT = '%Y-%m-%d'


# time = datetime.strptime("2021-02-13 00:00:00", FMT)

# date_time = datetime.now()

# print("now is {}".format(date_time.strftime(FMT)))

wb= openpyxl.load_workbook('customer order list.xlsx')
names = wb.sheetnames
print(wb.sheetnames)
sheet = wb.active

custom_order_list_dict = {}

for x in range (1,2):
    tmp_dict = {}
    for y in range(1,14):
        tmp_dict.update({y:sheet.cell(row=x,column=y).value})
    custom_order_list_dict.update({'row_1':tmp_dict})
pp = pprint.PrettyPrinter(indent=4)
pp.pprint(custom_order_list_dict)

for key, value in custom_order_list_dict.items():
    for k1, v1 in value.items():
        print("{} {} and typpe is {}".format(k1, v1, type(k1)))

for j in range(1, 14):
    i = j
    print("value is {}".format(custom_order_list_dict['row_1'][i]))

# iterate through excel and display data
print("total row is {}".format(sheet.max_row))
for line in range(2, sheet.max_row+1):
# for line in range(2, 20):
    tmp_dict = {}
    for colum_n in range(1, sheet.max_column+1):
        if isinstance(sheet.cell(row=line, column=colum_n).value, datetime):
            tmp_dict.update(
                {custom_order_list_dict['row_1'][colum_n] : 
                sheet.cell(row=line, column=colum_n).value.strftime(FMT)})
        else:

            tmp_dict.update({
                custom_order_list_dict['row_1'][colum_n] : 
                sheet.cell(row=line, column=colum_n).value})

    row_str = 'row_' + str(line)
    custom_order_list_dict.update({row_str:tmp_dict}) 

# pp.pprint(custom_order_list_dict)
out_file = open("customer_order_list.json", "w")
  
json.dump(custom_order_list_dict, out_file, indent = 4)
  
out_file.close()