import openpyxl
from datetime import datetime, timedelta
import pprint
import json

FMT = '%Y-%m-%d'


# time = datetime.strptime("2021-02-13 00:00:00", FMT)

# date_time = datetime.now()

# print("now is {}".format(date_time.strftime(FMT)))
file_dict = {'BOM list.xlsx' : 1,
            'Item List.xlsx' : 2,
            'customer order list.xlsx': 1}

for file_name, start_row in file_dict.items():
    wb= openpyxl.load_workbook(file_name)
    names = wb.sheetnames
    print(wb.sheetnames)
    sheet = wb.active

    data_dict = {}
    field_name_dict = {}

    for x in range (start_row,start_row + 1):

        for y in range(1,sheet.max_column + 1):
            field_name_dict.update({y:sheet.cell(row=x,column=y).value})
        # data_dict.update({'row_1':field_name_dict})
    pp = pprint.PrettyPrinter(indent=4)
    pp.pprint(field_name_dict)



    # # iterate through excel and display data
    print("{}'s total row is {}".format(file_name, sheet.max_row))
    for line in range(start_row + 1, sheet.max_row+1):
    # for line in range(2, 20):
        tmp_dict = {}
        for colum_n in range(1, sheet.max_column+1):
            if isinstance(sheet.cell(row=line, column=colum_n).value, datetime):
                tmp_dict.update(
                    {field_name_dict[colum_n] : 
                    sheet.cell(row=line, column=colum_n).value.strftime(FMT)})
            else:

                tmp_dict.update({
                    field_name_dict[colum_n] : 
                    sheet.cell(row=line, column=colum_n).value})

        row_str = 'row_' + str(line)
        data_dict.update({row_str:tmp_dict}) 

    # pp.pprint(custom_order_list_dict)
    with open(file_name.replace('.xlsx', '.json'), "w") as json_file:
        json.dump(data_dict, json_file, indent = 4)

all_dict = {'BOM list.xlsx' : {},
            'Item List.xlsx' : {},
            'customer order list.xlsx': {}}}

# Opening JSON file
for file_name, start_row in file_dict.items():
with open(file_name.replace('.xlsx', '.json')) as json_file:
    all_dict[file_name] = json.load(json_file)
