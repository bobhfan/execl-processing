import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import PatternFill, Alignment
from datetime import datetime, timedelta
from datetime import date as date_op
import pprint
import argparse
import json
from collections import OrderedDict, defaultdict
import holidays
import sys
import os



def get_file_list(dir, type):
    file_list = []
    data_files = [(x[0], x[2]) for x in os.walk(dir)]
    print(data_files)
    for path_files in data_files:
        for file_name in path_files[1]:
            if type in file_name and 'Zone' not in file_name:
                os.rename(dir+file_name, dir+file_name.lower())
                file_list.append(file_name.lower())
    return set(file_list)


def read_forcast_json(file_name):

    # FMT = '%d-%b-%Y'
    FMT = '%Y-%m-%d'
    YEAR_FMT = '%Y'

    print('handle {}'.format(file_name))

    if 'price chopper' in file_name:
        print('got it')

    current_date = datetime.now().date()
    previuos_year_date = date_op(current_date.year - 1, current_date.month, current_date.day)
    next_year_date = date_op(current_date.year + 1, current_date.month, current_date.day)


    pre_year = str(current_date.year - 1)
    cur_year = str(current_date.year )
    
    # date = datetime.now()


    wb= openpyxl.load_workbook(file_name, data_only=True)
    sheet = wb.active

    # for line in range(1, sheet.max_row+1):
    # for line in range(3, 4):
    #     for colum_n in range(1, sheet.max_column+1):
    #         print("{} and {}".format(sheet.cell(row=line, column=colum_n).value, type(sheet.cell(row=line, column=colum_n).value)))

    start_row = 3
    field_name_part_1_dict = {}
    field_name_dict = {}

    customizd_max_column = 73
    first_round = True

    skip_row = 0
    first_detection_on_row_3 = True
    found_item = False

    append_str = pre_year
    prev_column_name = ''

    for x in range (start_row,start_row + 3):
        if found_item:
            skip_row = x -1 - start_row
            break

        for y in range(1, customizd_max_column):
            # print("field {} at {}:{}".format(sheet.cell(row=x,column=y).value, start_row, y))

            if sheet.cell(row=x,column=1).value and 'Item' in sheet.cell(row=x,column=1).value:
                found_item = True
            elif not sheet.cell(row=x,column=1).value:
                break

            if not sheet.cell(row=x,column=y).value and y < 3:
                field_name = ''
                
            elif sheet.cell(row=x,column=y).value:
                field_name = sheet.cell(row=x,column=y).value
                
            if  'dec' in prev_column_name.lower() and 'jan' in field_name.lower():
                append_str = cur_year

            if y > 4:
                field_output_name = append_str + '\n' + field_name
            else:
                field_output_name = field_name

            prev_column_name = field_name
            field_name_part_1_dict.update({y:field_output_name})

    start_row += skip_row

    for x in range (start_row + 1, start_row + 2 ):
        for y in range(1, customizd_max_column):
            if sheet.cell(row=x,column=y).value:
                part_1 = field_name_part_1_dict[y]
                field_name_dict[y] = part_1 + '\n' + sheet.cell(row=x,column=y).value
            else:
                field_name_dict[y] = field_name_part_1_dict[y]


    print("Ther are {} rows in {} sheet of {} ".format(sheet.max_row, wb.sheetnames[0], file_name ))
    seq_no = 1
    data_dict = {}


    # not read the max_row of the file, 
    for line in range(start_row + 2, sheet.max_row+1):
    # for line in range(7, 8):
        tmp_dict = {}
        item_no = ''
        total_2021_not_null_cnt = 0
        for colum_n in range(1, customizd_max_column):
            # This is for extract item list info to create the dict for the first step
            # {item_no : { interesting_field : value}}
            if field_name_dict[colum_n]:
                if 'Item &' == field_name_dict[colum_n] or 'Item No.' == field_name_dict[colum_n]:
                    if isinstance(sheet.cell(row=line, column=colum_n).value, float):
                        item_no = str(int(sheet.cell(row=line, column=colum_n).value))
                    else:
                        item_no = sheet.cell(row=line, column=colum_n).value

                else:
                    if isinstance(sheet.cell(row=line, column=colum_n).value, datetime):
                        date_value = sheet.cell(row=line, column=colum_n).value.strftime(FMT)
                        tmp_dict.update({field_name_dict[colum_n] : date_value})
                    else:
                        if '2021' in field_name_dict[colum_n] and sheet.cell(row=line, column=colum_n).value:
                            total_2021_not_null_cnt += 1
                        tmp_dict.update({field_name_dict[colum_n] : sheet.cell(row=line, column=colum_n).value})
        if item_no and total_2021_not_null_cnt > 1:
            data_dict[item_no] = tmp_dict


    print("Wrote {} records into the temporary file ".format(len(data_dict)))
    if len(data_dict) == 0:
        print('something error')

    if 'price chopper' in file_name:
        print('got it')
    data_dict_tmp_2 = {}
    for key, top_dict in data_dict.items():
        data_dict_tmp_2[key] = {}
        data_dict_tmp_2[key]['Description'] = top_dict['Description']

        for field, value in top_dict.items():
            if len(field.splitlines()) < 3:
                continue
            line_1, line_2, line_3 = field.splitlines()
            if cur_year in line_1 and ('Actual' in line_3 or '2021' in line_3):
                month_key = '{}{}'.format(line_1, line_2)
                data_dict_tmp_2[key][month_key] = value

    with open(file_name.replace('.xlsx', '_step_1.json'), "w") as json_file:
        json.dump(data_dict_tmp_2, json_file, indent = 4)

def read_forcast(dir, bom_json_file_name):
    xlsx_files = get_file_list(dir, '.xlsx')
    for xlsx_file in xlsx_files:
        read_forcast_json(dir + xlsx_file)
    
    total_forcast_dict = {}
    json_files = get_file_list(dir, '.json')
    for json_file in json_files:
        if 'total_forcast' in json_file:
            continue
        with open(dir + json_file) as tmp_file:
            tmp_dict = json.load(tmp_file)
            for first_key, value in tmp_dict.items():
                print('now handle {}'.format(first_key))
                second_key = str(value['Description'])
                if first_key in total_forcast_dict:
                    print("repeat item no in different forcast {} in {}".format(first_key, json_file))
                    total_forcast_dict[first_key][second_key] = value
                else:
                    total_forcast_dict[first_key] = {}
                    total_forcast_dict[first_key][second_key] = value


    with open(dir + 'total_forcast_step_1.json', "w") as json_file:
        json.dump(total_forcast_dict, json_file, indent = 4)   

    with open(bom_json_file_name) as json_read_file:
        bom_dict = json.load(json_read_file)


    currrent_year = str(datetime.now().date().year)
    prev_year = str(datetime.now().date().year - 1)
    
    forcast_material_tmp_dict = {}
    for item_no, top_dict in total_forcast_dict.items():
        if item_no not in bom_dict:
            print("forcasdt error: item no {} didn't exist in bom".format(item_no))
            continue
        for description, each_month_dict in top_dict.items():
            for month, value in each_month_dict.items():
                if 'Description' in month:
                    continue

                material_dict = bom_dict[item_no]
                for materail_no, quantity_dict in material_dict.items():
                    if quantity_dict['Quantity'] \
                        and quantity_dict['Scrap %'] \
                        and each_month_dict[month] \
                        and isinstance(each_month_dict[month], int): 
                        quantity = quantity_dict['Quantity']
                        scrap = quantity_dict['Scrap %']
                        material_quantity = float(each_month_dict[month]) * float(quantity) * (1.0 + 0.01 * scrap)
                        combined_key = "{}----{}----{}----{}".format(materail_no, month.upper(), item_no, description)
                        forcast_material_tmp_dict[combined_key] = material_quantity
                    else:
                        break
    
    with open(dir + 'total_forcast_step_2.json', "w") as json_file:   
        json.dump(forcast_material_tmp_dict, json_file, indent = 4)   

    forcast_material_tmp_dict_2 = {}

    for key, value in forcast_material_tmp_dict.items():
        item_no, month, order_no, descr = key.split('----')
        if item_no not in forcast_material_tmp_dict_2:
            forcast_material_tmp_dict_2[item_no] = {}
        if month not in forcast_material_tmp_dict_2[item_no]:
            forcast_material_tmp_dict_2[item_no][month] = {}
        forcast_material_tmp_dict_2[item_no][month][order_no] = value

    with open(dir + 'total_forcast_step_3.json', "w") as json_file:
        json.dump(forcast_material_tmp_dict_2, json_file, indent = 4)   

    forcast_material_tmp_dict_3 = {}
    for item_no, month_dict in forcast_material_tmp_dict_2.items():
        if 'RF 007' in item_no:
            print('got it')
        if item_no not in forcast_material_tmp_dict_3:
            forcast_material_tmp_dict_3[item_no] = {}
        for month, order_quantity_dict in month_dict.items():
            sum_quantity = 0
            for quantity in order_quantity_dict.values():
                sum_quantity += quantity
            
            forcast_material_tmp_dict_3[item_no][month] = sum_quantity

    with open(dir + 'total_forcast_step_4.json', "w") as json_file:
        json.dump(forcast_material_tmp_dict_3, json_file, indent = 4)       
        