import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import PatternFill, Alignment
from datetime import datetime, timedelta
from datetime import date as date_op
import pprint
import argparse
import json
from collections import OrderedDict, defaultdict
import holidays
import sys
import os


    


FMT = '%d-%b-%Y'
FMT_1 = '%Y-%m-%d, %H:%M:%S'

time = date_op(2021, 1, 1)

# date_time = datetime.now()

# print("now is {}".format(date_time.strftime(FMT)))

def read_film_amend_data(file_name, interest_field_list):


# date = datetime.now()

    wb= openpyxl.load_workbook(file_name)
    print(wb.sheetnames)
    sheet = wb.active

    # for line in range(1, sheet.max_row+1):
    for line in range(6, 7):
        for colum_n in range(1, sheet.max_column+1):
            print("{} and {}".format(sheet.cell(row=line, column=colum_n).value, type(sheet.cell(row=line, column=colum_n).value)))

    start_row = 6
    field_name_dict = {}

    for x in range (start_row,start_row + 1):
        for y in range(1,sheet.max_column + 1):
            field_name_dict.update({y:sheet.cell(row=x,column=y).value})

    print("Ther are {} rows in {} sheet of {} ".format(sheet.max_row, wb.sheetnames[0], file_name ))
    seq_no = 1
    data_dict = {}
    for line in range(start_row + 1, sheet.max_row+1):
    # for line in range(3, 155):
        tmp_dict = {}
        for colum_n in range(1, sheet.max_column+1):
            # This is for extract item list info to create the dict for the first step
            # {item_no : { interesting_field : value}}
            if field_name_dict[colum_n]:
                if 'Item' in field_name_dict[colum_n]:
                    if isinstance(sheet.cell(row=line, column=colum_n).value, float):
                        item_no = str(int(sheet.cell(row=line, column=colum_n).value))
                    else:
                        item_no = sheet.cell(row=line, column=colum_n).value

                elif field_name_dict[colum_n] in interest_field_list:
                    if isinstance(sheet.cell(row=line, column=colum_n).value, datetime):
                        date_value = sheet.cell(row=line, column=colum_n).value.strftime(FMT)
                        tmp_dict.update({field_name_dict[colum_n] : date_value})
                    else:
                        tmp_dict.update({field_name_dict[colum_n] : sheet.cell(row=line, column=colum_n).value})
        
        data_dict[item_no] = tmp_dict

    return data_dict