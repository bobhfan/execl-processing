import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import PatternFill, Alignment
from datetime import datetime, timedelta
from datetime import date as date_op
import pprint
import argparse
import json
from collections import OrderedDict, defaultdict
import holidays
import sys
import os


# FMT = '%d-%b-%Y'
FMT = '%Y-%m-%d'

time = date_op(2021, 1, 1)

# date_time = datetime.now()

# print("now is {}".format(date_time.strftime(FMT)))


# date = datetime.now()

file_name = 'Canadian Tire May 2021.xlsx'

wb= openpyxl.load_workbook(file_name, data_only=True)
print(wb.sheetnames)
sheet = wb.active

# for line in range(1, sheet.max_row+1):
# for line in range(3, 4):
#     for colum_n in range(1, sheet.max_column+1):
#         print("{} and {}".format(sheet.cell(row=line, column=colum_n).value, type(sheet.cell(row=line, column=colum_n).value)))

start_row = 3
field_name_dict = {}

for x in range (start_row,start_row + 1):
    for y in range(1,sheet.max_column + 1):
        field_name_dict.update({y:sheet.cell(row=x,column=y).value})

for x in range (5, 6):
    for y in range(1,sheet.max_column + 1):
        if sheet.cell(row=x,column=y).value:
            print(sheet.cell(row=x,column=y).value, sep='*', end=" ")

print("Ther are {} rows in {} sheet of {} ".format(sheet.max_row, wb.sheetnames[0], file_name ))
seq_no = 1
data_dict = {}
for line in range(start_row + 1, sheet.max_row+1):
# for line in range(3, 155):
    tmp_dict = {}
    for colum_n in range(1, sheet.max_column+1):
        # This is for extract item list info to create the dict for the first step
        # {item_no : { interesting_field : value}}
        if field_name_dict[colum_n]:
            if 'Item &' == field_name_dict[colum_n]:
                if isinstance(sheet.cell(row=line, column=colum_n).value, float):
                    item_no = str(int(sheet.cell(row=line, column=colum_n).value))
                else:
                    item_no = sheet.cell(row=line, column=colum_n).value

            else:
                if isinstance(sheet.cell(row=line, column=colum_n).value, datetime):
                    date_value = sheet.cell(row=line, column=colum_n).value.strftime(FMT)
                    tmp_dict.update({field_name_dict[colum_n] : date_value})
                else:
                    tmp_dict.update({field_name_dict[colum_n] : sheet.cell(row=line, column=colum_n).value})
    
    data_dict[item_no] = tmp_dict

        # pp.pprint(custom_order_list_dict)
with open(file_name.replace('.xlsx', '_step_1.json'), "w") as json_file:
    json.dump(data_dict, json_file, indent = 4)